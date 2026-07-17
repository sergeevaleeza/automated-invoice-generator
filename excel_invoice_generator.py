#!/usr/bin/env python3
"""
Excel Invoice Generator
Produces a print-ready .xlsx patient invoice matching the approved layout in
tests/fixtures/Example_2026_Invoice_07162026.xlsx exactly. Consumes the same
inputs PatientInvoiceGenerator._generate_pdf_invoice() consumes (PatientData,
InvoiceLine list, total_due, the per-patient DataFrame, statement/payment-due
dates, and the has_cpt flag) so no billing business logic is duplicated here
— this module is presentation only.
"""
import math
from io import BytesIO
from pathlib import Path
from datetime import datetime
from typing import List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from invoice_models import PatientData, InvoiceLine, format_date_for_display
from clinic_config import load_clinic_config
from qr_code import generate_qr_png_bytes, qr_settings

# --- Static layout config: labels, widths, margins, row heights, fonts ----
# Clinic identity (name, addresses, EIN/NPI, payment info) is NOT here — it's
# loaded at call time from clinic_config.json via _clinic_derived_config()
# below, so no real practice's business identity is hardcoded in source.
LAYOUT_CONFIG = {
    "statement_title": "PATIENT STATEMENT",
    "table_headers": ["Service Date(s)", "Description", "Amount Paid", "Copay/Deductible"],
    "amount_due_label": "YOUR PORTION DUE:",
    "amount_enclosed_label": "AMOUNT ENCLOSED:",
    "default_service_description": "Psychotherapy and/or Med Management",

    "font_family": "Arial",
    # Pixel targets (per on-screen review): A=105, B=64, C=225, D=140, E=140.
    # Converted via width = pixels / 7, confirmed against measured on-screen
    # widths (an earlier width = (px-5)/7 guess rendered 5px short on every
    # single column). D/E are equal so the payment-notice box has enough
    # room for its text to wrap the same way it was authored.
    "col_widths": {"A": 15.0, "B": 9.14, "C": 32.14, "D": 20.0, "E": 20.0},
    "header_fill_color": "D9D9D9",

    "margins": dict(left=0.65, right=0.65, top=0.4, bottom=0.6, header=0.2, footer=0.2),

    "fonts": dict(
        clinic_header=12, contact=10, title=13, statement_label=10,
        table_header=10, item=9, subtotal_total=10, signature=11,
    ),
    "row_heights": dict(
        clinic_header=18, spacer_after_clinic=6, contact=15.95, spacer_after_contact=12,
        info_box=15, spacer_after_info=12, title=20.1, spacer_after_title=6,
        statement=15, spacer_after_statement=12, item_header=18, item=15,
        spacer_after_items=12, amount_box=17.1, spacer_after_amount=18, signature=15,
    ),
}

_LETTER_HEIGHT_IN = 11.0
LAYOUT_CONFIG["usable_height_pt"] = (
    _LETTER_HEIGHT_IN - LAYOUT_CONFIG["margins"]["top"] - LAYOUT_CONFIG["margins"]["bottom"]
) * 72


def _clinic_derived_config(clinic: dict) -> dict:
    """Build the clinic-identity-derived display strings (header lines,
    payment-notice text, signature label, footer) from a loaded clinic
    config dict. Wording/formatting is fixed (matches the approved
    layout); only the underlying values come from clinic_config.json."""
    return {
        "clinic_name": clinic["clinic_name"],
        "doctor_name": clinic["doctor_name"],
        "specialty": clinic["specialty"],
        "ein_npi": f"EIN: {clinic['ein']}    NPI: {clinic['npi']}",
        "office_address": f"OFFICE ADDRESS: {clinic['office_address']}",
        "mailing_address": f"MAILING ADDRESS: {clinic['mailing_address']}",
        "email": f"EMAIL: {clinic['email']}",
        "website": f"WEBSITE: {clinic['website']}",
        # Fixed literal lines (not auto-wrapped) — sized to fit exactly the 5
        # rows of the notice box (rows 10-14) at 15.0pt each.
        "payment_instructions": [
            "Please note we do not accept credit cards.",
            f"1. Zelle {clinic['zelle_email']} ",
            "   (IRA Billing and Mgmt)",
            f"2. Check payable to: {clinic['check_payable_to']}",
            f"    {clinic['mailing_address']}",
        ],
        "signature_label": f"Provider Signature - {clinic['provider_name_for_signature']}",
        "footer_line1": f"If you have questions regarding your bill, please contact us at {clinic['phone']}.",
        "footer_line2": f"For current pricing, please visit: {clinic['pricing_page_url']}",
        "show_qr": qr_settings(clinic)[0],
        "qr_content": qr_settings(clinic)[1],
    }

# Matches the fixture's literal number_format string exactly (backslash-escaped).
CURRENCY_FMT = r'\$#,##0.00;\-\$#,##0.00;"$ -"'
THIN = Side(style="thin", color="000000")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _clean_postal_code(postal_code: str) -> str:
    if postal_code and '.' in postal_code:
        return postal_code.split('.')[0]
    return postal_code


def _estimate_chars_per_line(col_width_units: float, font_size: int) -> int:
    """Estimate how many characters fit on one wrapped line inside a merged
    range of the given total column width (Excel width units) and font size.
    An Excel width unit is roughly one character of the workbook's default
    ~11pt font; scale for the actual font size and subtract a small margin
    for cell padding."""
    base_chars = max(col_width_units - 2, 1)
    scale = 11.0 / max(font_size, 6)
    return max(int(base_chars * scale), 8)


def _count_wrapped_lines(text_lines: List[str], col_width_units: float, font_size: int) -> int:
    """Estimate the total visual (wrapped) line count for a list of logical
    text lines once Excel wraps them inside a merged cell of the given width.
    Excel does not auto-grow merged-cell row heights for wrapped text, so
    row spans must be sized from this estimate rather than the raw line
    count, or trailing lines get visually clipped."""
    chars_per_line = _estimate_chars_per_line(col_width_units, font_size)
    total = 0
    for line in text_lines:
        if not line:
            total += 1
            continue
        total += max(1, math.ceil(len(line) / chars_per_line))
    return total


def apply_box_border(ws: Worksheet, min_row: int, min_col: int, max_row: int, max_col: int,
                      style: str = "thin") -> None:
    """Draw a single outline box around a cell range (merged or not, single-
    or multi-column), setting only the true perimeter edge on each boundary
    cell — no internal divider lines. For a single-column range, the one
    column correctly gets both its left and right edge styled."""
    side = Side(style=style, color="000000")
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            existing = cell.border
            cell.border = Border(
                left=side if c == min_col else existing.left,
                right=side if c == max_col else existing.right,
                top=side if r == min_row else existing.top,
                bottom=side if r == max_row else existing.bottom,
            )


def _build_workbook(patient: PatientData, total_due: float, patient_df: pd.DataFrame,
                     statement_date: datetime, payment_due_date: datetime,
                     has_cpt: bool, cfg: dict) -> Tuple[Workbook, Worksheet, int, int]:
    """Build the invoice worksheet. Returns (workbook, worksheet, header_row_idx, last_row)."""
    fonts = cfg["fonts"]
    heights = cfg["row_heights"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    row = 1

    def set_row_height(r: int, h: float):
        ws.row_dimensions[r].height = h

    def merged(r1, c1, r2, c2, value, font, alignment):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell = ws.cell(row=r1, column=c1, value=value)
        cell.font = font
        cell.alignment = alignment
        return cell

    clinic_font = Font(name=cfg["font_family"], size=fonts["clinic_header"], bold=True)
    contact_font = Font(name=cfg["font_family"], size=fonts["contact"], bold=True)
    title_font = Font(name=cfg["font_family"], size=fonts["title"], bold=True)
    statement_font = Font(name=cfg["font_family"], size=fonts["statement_label"], bold=True)
    table_header_font = Font(name=cfg["font_family"], size=fonts["table_header"], bold=True)
    item_font = Font(name=cfg["font_family"], size=fonts["item"])
    subtotal_total_font = Font(name=cfg["font_family"], size=fonts["subtotal_total"], bold=True)
    sig_font = Font(name=cfg["font_family"], size=fonts["signature"], bold=True)

    center = Alignment(horizontal="center", vertical="center")
    left_top_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    left_center = Alignment(horizontal="left", vertical="center")
    right_center = Alignment(horizontal="right", vertical="center")

    # --- Clinic header block (rows 1-3, + optional EIN/NPI row) ---
    for text in (cfg["clinic_name"], cfg["doctor_name"], cfg["specialty"]):
        merged(row, 1, row, 5, text, clinic_font, center)
        set_row_height(row, heights["clinic_header"])
        row += 1
    if has_cpt:
        merged(row, 1, row, 5, cfg["ein_npi"], clinic_font, center)
        set_row_height(row, heights["clinic_header"])
        row += 1
    set_row_height(row, heights["spacer_after_clinic"])
    row += 1

    # --- Contact info block ---
    for text in (cfg["office_address"], cfg["mailing_address"], cfg["email"], cfg["website"]):
        merged(row, 1, row, 5, text, contact_font, center)
        set_row_height(row, heights["contact"])
        row += 1
    set_row_height(row, heights["spacer_after_contact"])
    row += 1

    # --- Patient address block (A:B) + payment notice box (D:E), column C is a spacer ---
    display_postal = _clean_postal_code(patient.postal_code)
    patient_lines = [f"{patient.first_name.upper()} {patient.last_name.upper()}"]
    if patient.address_line1:
        patient_lines.append(patient.address_line1)
    if patient.address_line2:
        patient_lines.append(patient.address_line2)
    patient_lines.append(f"{patient.city}, {patient.state} {display_postal}")
    patient_text = "\n".join(patient_lines)
    payment_text = "\n".join(cfg["payment_instructions"])

    # The payment-notice text is fixed literal content designed to fit
    # exactly 5 rows (see _clinic_derived_config) — only the patient address varies
    # per invoice, so only it needs dynamic wrap-based sizing. The box
    # floors at 5 rows to match the fixture and grows only for addresses
    # too long to fit that (Excel won't auto-grow wrapped merged-cell rows).
    wrapped_patient_lines = _count_wrapped_lines(
        patient_lines, cfg["col_widths"]["A"] + cfg["col_widths"]["B"], fonts["contact"])

    info_row_start = row
    n_info_lines = max(wrapped_patient_lines, 5)
    info_row_end = info_row_start + n_info_lines - 1

    merged(info_row_start, 1, info_row_end, 2, patient_text, contact_font, left_top_wrap)
    merged(info_row_start, 4, info_row_end, 5, payment_text, contact_font, left_top_wrap)
    # No border on the payment-notice box — matches the approved fixture exactly.
    for r in range(info_row_start, info_row_end + 1):
        set_row_height(r, heights["info_box"])

    if cfg.get("show_qr") and cfg.get("qr_content"):
        # Column C is a deliberate spacer between the patient-address and
        # payment-notice boxes — always blank, regardless of how many rows
        # the boxes span — so a floating image anchored there can't overlap
        # existing text or disturb the tested grid/merge/print-area layout.
        qr_buf = BytesIO(generate_qr_png_bytes(cfg["qr_content"]))
        qr_image = XLImage(qr_buf)
        # openpyxl sizes images in pixels at a 96dpi assumption when
        # converting to the saved anchor's EMU extent (verified empirically:
        # 65px round-tripped to 0.677in, not the 0.9in intended) — 86px/96dpi ≈ 0.9in.
        qr_image.width = qr_image.height = 86
        qr_image.anchor = f"C{info_row_start}"
        ws.add_image(qr_image)

    row = info_row_end + 1
    set_row_height(row, heights["spacer_after_info"])
    row += 1

    # --- Title ---
    merged(row, 1, row, 5, cfg["statement_title"], title_font, center)
    set_row_height(row, heights["title"])
    row += 1
    set_row_height(row, heights["spacer_after_title"])
    row += 1

    # --- Statement date row (labels at C, values merged D:E) ---
    ws.cell(row=row, column=3, value="STATEMENT DATE:").font = statement_font
    ws.cell(row=row, column=3).alignment = left_center
    merged(row, 4, row, 5, "Payment due date:", statement_font, left_center)
    set_row_height(row, heights["statement"])
    row += 1
    ws.cell(row=row, column=3, value=statement_date.strftime('%m/%d/%Y')).font = statement_font
    ws.cell(row=row, column=3).alignment = left_center
    merged(row, 4, row, 5, payment_due_date.strftime('%m/%d/%Y'), statement_font, left_center)
    set_row_height(row, heights["statement"])
    row += 1
    set_row_height(row, heights["spacer_after_statement"])
    row += 1

    # --- Line-item table: A=date, B:C merged=description, D=amount paid, E=copay ---
    header_row_idx = row
    header_cols = [(1, 1, cfg["table_headers"][0]), (2, 3, cfg["table_headers"][1]),
                   (4, 4, cfg["table_headers"][2]), (5, 5, cfg["table_headers"][3])]
    for c1, c2, text in header_cols:
        if c1 == c2:
            cell = ws.cell(row=row, column=c1, value=text)
        else:
            cell = merged(row, c1, row, c2, text, table_header_font, left_center)
        cell.font = table_header_font
        cell.alignment = Alignment(horizontal="left" if c1 <= 2 else "right", vertical="center")
        for c in range(c1, c2 + 1):
            ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=cfg["header_fill_color"])
            ws.cell(row=row, column=c).border = CELL_BORDER
    set_row_height(row, heights["item_header"])
    row += 1

    def add_item_row(date_val, desc, paid_val, copay_val, bold=False, top_border_style="thin"):
        nonlocal row
        font = subtotal_total_font if bold else item_font
        border = Border(left=THIN, right=THIN, bottom=THIN, top=Side(style=top_border_style, color="000000"))
        c_date = ws.cell(row=row, column=1, value=date_val)
        merged(row, 2, row, 3, desc, font, left_center)
        c_paid = ws.cell(row=row, column=4, value=paid_val)
        c_copay = ws.cell(row=row, column=5, value=copay_val)
        for c in range(1, 6):
            cell = ws.cell(row=row, column=c)
            cell.font = font
            cell.border = border
        c_date.alignment = left_center
        ws.cell(row=row, column=2).alignment = left_center
        c_paid.alignment = right_center
        c_copay.alignment = right_center
        if paid_val is not None:
            c_paid.number_format = CURRENCY_FMT
        if copay_val is not None:
            c_copay.number_format = CURRENCY_FMT
        set_row_height(row, heights["item"])
        row += 1

    previous_balance = float(patient_df.iloc[0].get('previous_balance', 0))
    total_paid_display = 0.0
    total_copay_display = 0.0

    if previous_balance > 0:
        add_item_row("", "Previous Balance", 0.0, previous_balance)
        total_copay_display += previous_balance
    elif previous_balance < 0:
        add_item_row("", "Previous Balance (Overpaid)", abs(previous_balance), 0.0)
        total_paid_display += abs(previous_balance)

    for _, prow in patient_df.iterrows():
        paid_amount = float(prow.get('paid', 0))
        copay_amount = float(prow.get('copay', 0))
        if paid_amount > 0 or copay_amount > 0:
            display_date = format_date_for_display(prow['visit_date'])
            service_type = str(prow.get('type_of_service', '')).strip() or cfg["default_service_description"]
            add_item_row(display_date, service_type, paid_amount, copay_amount)
            total_paid_display += paid_amount
            total_copay_display += copay_amount

    add_item_row("", "SUBTOTAL", total_paid_display, total_copay_display, bold=True, top_border_style="medium")
    add_item_row("", "TOTAL", None, total_due, bold=True, top_border_style="double")

    set_row_height(row, heights["spacer_after_items"])
    row += 1

    # --- Bottom boxes: "YOUR PORTION DUE" (col C) and "AMOUNT ENCLOSED"
    # (D:E merged) — both get a complete outline border. ---
    portion_label_row = row
    ws.cell(row=row, column=3, value=cfg["amount_due_label"]).font = subtotal_total_font
    ws.cell(row=row, column=3).alignment = left_center
    merged(row, 4, row, 5, cfg["amount_enclosed_label"], subtotal_total_font, left_center)
    set_row_height(row, heights["amount_box"])
    row += 1
    portion_value_row = row
    due_cell = ws.cell(row=row, column=3, value=total_due)
    due_cell.font = subtotal_total_font
    due_cell.alignment = left_center
    due_cell.number_format = CURRENCY_FMT
    merged(row, 4, row, 5, "", subtotal_total_font, left_center)
    set_row_height(row, heights["amount_box"])
    row += 1
    apply_box_border(ws, portion_label_row, 3, portion_value_row, 3)
    apply_box_border(ws, portion_label_row, 4, portion_value_row, 5)
    set_row_height(row, heights["spacer_after_amount"])
    row += 1

    # --- Provider signature ---
    merged(row, 1, row, 5, "_________________________________", sig_font, right_center)
    set_row_height(row, heights["signature"])
    row += 1
    merged(row, 1, row, 5, cfg["signature_label"], sig_font, right_center)
    set_row_height(row, heights["signature"])
    row += 1

    last_row = row - 1
    return wb, ws, header_row_idx, last_row


def _apply_page_setup(ws: Worksheet, last_row: int, header_row_idx: int, cfg: dict):
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(**cfg["margins"])
    ws.print_area = f"A1:E{last_row}"
    ws.print_title_rows = f"{header_row_idx}:{header_row_idx}"
    ws.sheet_view.showGridLines = False

    for col_letter, width in cfg["col_widths"].items():
        ws.column_dimensions[col_letter].width = width

    ws.oddFooter.center.text = f"{cfg['footer_line1']}\n{cfg['footer_line2']}"
    ws.oddFooter.center.size = 8
    ws.oddFooter.center.font = "Arial,Regular"


def generate_excel_invoice(patient: PatientData, lines: List[InvoiceLine], total_due: float,
                            patient_df: pd.DataFrame, statement_date: datetime,
                            payment_due_date: datetime, has_cpt: bool, output_path: Path,
                            clinic: Optional[dict] = None) -> None:
    """Generate a print-ready Excel invoice matching the approved fixture layout.

    Takes the same inputs already assembled for _generate_pdf_invoice (lines,
    total_due, patient_df, has_cpt) — no billing logic is recomputed here.
    Rows extend downward for item counts beyond the fixture's 8, keeping the
    same relative structure; fitToPage print scaling handles fitting the
    result onto one printed Letter page regardless of row count.

    clinic: pre-loaded clinic identity dict (see clinic_config.py). Defaults
    to load_clinic_config() — callers that already loaded it once (e.g. to
    share across the PDF and Excel generators in one batch run) should pass
    it through instead of reloading from disk per invoice. Tests inject a
    fixed dict here directly rather than depending on the real, gitignored
    clinic_config.json.
    """
    clinic = clinic if clinic is not None else load_clinic_config()
    cfg = {**LAYOUT_CONFIG, **_clinic_derived_config(clinic)}

    wb, ws, header_row_idx, last_row = _build_workbook(
        patient, total_due, patient_df, statement_date, payment_due_date, has_cpt, cfg
    )
    _apply_page_setup(ws, last_row, header_row_idx, cfg)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
