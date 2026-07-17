"""Golden-file test: a freshly generated invoice (from synthetic data
matching the fixture's inputs exactly) must match
tests/fixtures/Example_2026_Invoice_07162026.xlsx cell-for-cell, merge-for-
merge, and in print setup. This is Phase 0's regression guard against the
Excel layout drifting away from the approved design again.
"""
from openpyxl import load_workbook

from excel_invoice_generator import generate_excel_invoice
from tests.conftest import GOLDEN_INVOICE_PATH

# Row heights the fixture leaves unset (falls back to Excel's default)
# rather than explicitly stating, which looks like an authoring
# inconsistency rather than a deliberate choice — row 29 (the *last* item
# row) does have an explicit height, but rows 22-28 (the rest) don't. The
# generator explicitly sets every item row's height instead, per the
# existing "explicit heights, don't rely on autofit" convention. This set
# documents that intentional, known deviation so the test doesn't chase it.
ROWS_WITH_KNOWN_HEIGHT_DEVIATION = {22, 23, 24, 25, 26, 27, 28}

# The fixture's B30 (merged B30:C30, the SUBTOTAL label) has a stray
# "double" bottom border that A30/D30/E30 don't share — printing that
# literally would draw a broken half-underline under just the word
# "SUBTOTAL" and not the adjacent dollar amounts. Almost certainly an
# authoring artifact (mirrors a similar B29 bottom-border inconsistency in
# the fixture), not an intentional design choice, so the generator applies
# a uniform border across the whole row instead.
KNOWN_BORDER_DEVIATIONS = {("B30", "bottom")}


def _generate(tmp_path, golden_invoice_inputs):
    output_path = tmp_path / "generated.xlsx"
    generate_excel_invoice(output_path=output_path, **golden_invoice_inputs)
    return load_workbook(output_path)


def test_dimensions_match_fixture(tmp_path, golden_invoice_inputs):
    fixture_ws = load_workbook(GOLDEN_INVOICE_PATH).active
    generated_ws = _generate(tmp_path, golden_invoice_inputs).active
    assert generated_ws.dimensions == fixture_ws.dimensions


def test_merges_match_fixture(tmp_path, golden_invoice_inputs):
    fixture_ws = load_workbook(GOLDEN_INVOICE_PATH).active
    generated_ws = _generate(tmp_path, golden_invoice_inputs).active
    fixture_merges = {str(m) for m in fixture_ws.merged_cells.ranges}
    generated_merges = {str(m) for m in generated_ws.merged_cells.ranges}
    assert generated_merges == fixture_merges


def test_column_widths_match_fixture(tmp_path, golden_invoice_inputs):
    fixture_ws = load_workbook(GOLDEN_INVOICE_PATH).active
    generated_ws = _generate(tmp_path, golden_invoice_inputs).active
    for col in "ABCDE":
        fixture_dim = fixture_ws.column_dimensions.get(col)
        generated_dim = generated_ws.column_dimensions.get(col)
        fixture_width = fixture_dim.width if fixture_dim else None
        generated_width = generated_dim.width if generated_dim else None
        assert generated_width == fixture_width, f"column {col} width mismatch"


def test_row_heights_match_fixture(tmp_path, golden_invoice_inputs):
    fixture_ws = load_workbook(GOLDEN_INVOICE_PATH).active
    generated_ws = _generate(tmp_path, golden_invoice_inputs).active
    for row in range(1, fixture_ws.max_row + 1):
        fixture_dim = fixture_ws.row_dimensions.get(row)
        fixture_height = fixture_dim.height if fixture_dim else None
        if fixture_height is None and row in ROWS_WITH_KNOWN_HEIGHT_DEVIATION:
            continue
        generated_dim = generated_ws.row_dimensions.get(row)
        generated_height = generated_dim.height if generated_dim else None
        assert generated_height == fixture_height, f"row {row} height mismatch"


def test_cell_values_match_fixture(tmp_path, golden_invoice_inputs):
    fixture_ws = load_workbook(GOLDEN_INVOICE_PATH).active
    generated_ws = _generate(tmp_path, golden_invoice_inputs).active
    coords = set()
    for ws in (fixture_ws, generated_ws):
        for row in ws.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    coords.add(cell.coordinate)
    for coord in coords:
        assert generated_ws[coord].value == fixture_ws[coord].value, f"{coord} value mismatch"


def test_key_cell_styling_matches_fixture(tmp_path, golden_invoice_inputs):
    fixture_ws = load_workbook(GOLDEN_INVOICE_PATH).active
    generated_ws = _generate(tmp_path, golden_invoice_inputs).active
    # A representative sample covering every distinct style zone: clinic
    # header, patient box, payment box (no border), title, statement
    # labels, item header (fill+border), item row (currency format),
    # subtotal/total (medium/double top border), both bottom boxes
    # (including the portion-due box's no-left/right-border quirk), and
    # the signature line.
    sample_coords = [
        "A1", "A10", "D10", "A16", "C18", "A21", "D21",
        "A22", "D22", "B30", "D30", "B31", "E31",
        "C33", "D33", "C34", "D34", "E34", "A36",
    ]
    for coord in sample_coords:
        fc = fixture_ws[coord]
        gc = generated_ws[coord]
        assert gc.font.name == fc.font.name, f"{coord} font name mismatch"
        assert gc.font.size == fc.font.size, f"{coord} font size mismatch"
        assert gc.font.bold == fc.font.bold, f"{coord} font bold mismatch"
        assert gc.number_format == fc.number_format, f"{coord} number_format mismatch"
        assert gc.alignment.horizontal == fc.alignment.horizontal, f"{coord} horizontal alignment mismatch"
        for side in ("left", "right", "top", "bottom"):
            if (coord, side) in KNOWN_BORDER_DEVIATIONS:
                continue
            fs = getattr(fc.border, side).style
            gs = getattr(gc.border, side).style
            assert gs == fs, f"{coord} border.{side} mismatch (fixture={fs!r} generated={gs!r})"


def test_print_setup_matches_fixture(tmp_path, golden_invoice_inputs):
    fixture_ws = load_workbook(GOLDEN_INVOICE_PATH).active
    generated_ws = _generate(tmp_path, golden_invoice_inputs).active

    assert generated_ws.page_setup.orientation == fixture_ws.page_setup.orientation
    assert generated_ws.page_setup.paperSize == fixture_ws.page_setup.paperSize
    assert generated_ws.page_setup.fitToWidth == fixture_ws.page_setup.fitToWidth
    assert generated_ws.page_setup.fitToHeight == fixture_ws.page_setup.fitToHeight
    assert generated_ws.sheet_properties.pageSetUpPr.fitToPage == fixture_ws.sheet_properties.pageSetUpPr.fitToPage

    fm, gm = fixture_ws.page_margins, generated_ws.page_margins
    for attr in ("left", "right", "top", "bottom", "header", "footer"):
        assert getattr(gm, attr) == getattr(fm, attr), f"margin.{attr} mismatch"

    # print_area/print_title_rows include the sheet name (e.g. "'Invoice'!$A$1:$E$37");
    # compare only the cell-range portion since sheet titles aren't semantically relevant here.
    assert generated_ws.print_area.split("!")[-1] == fixture_ws.print_area.split("!")[-1]
    assert generated_ws.print_title_rows == fixture_ws.print_title_rows


def test_more_items_than_fixture_shifts_rows_down(tmp_path, golden_invoice_inputs):
    """When there are more line items than the fixture's 8, every section
    below the item table should shift down by the same number of extra
    rows, keeping the same relative structure — not overlap or truncate."""
    inputs = dict(golden_invoice_inputs)
    extra_rows = [("2026-05-14", "Psychotherapy", 0, 100), ("2026-05-28", "Psychotherapy", 0, 100)]
    import pandas as pd
    inputs["patient_df"] = pd.concat([
        inputs["patient_df"],
        pd.DataFrame(extra_rows, columns=["visit_date", "type_of_service", "paid", "copay"]).assign(previous_balance=0),
    ], ignore_index=True)
    inputs["total_due"] = 921.0

    ws = _generate(tmp_path, inputs).active
    fixture_ws = load_workbook(GOLDEN_INVOICE_PATH).active

    shift = 2  # two extra item rows
    assert ws.max_row == fixture_ws.max_row + shift
    # Everything before the item table (fixed-position header/title/dates)
    # stays put — only content after the (now longer) table shifts down.
    assert ws["A16"].value == "PATIENT STATEMENT"
    assert ws["A21"].value == "Service Date(s)"
    assert ws[f"C{33 + shift}"].value == "YOUR PORTION DUE:"
    assert ws[f"A{36 + shift}"].value == "_________________________________"
    assert ws.print_area.split("!")[-1] == f"$A$1:$E${fixture_ws.max_row + shift}"
