"""Tests for the tear-off receipt stub (STATEMENT DATE/Payment due date and
YOUR PORTION DUE/AMOUNT ENCLOSED moved above PATIENT STATEMENT, with a
dashed tear line) and the enlarged Zelle QR code, in both the PDF and
Excel generators. All patient data here is synthetic."""
import base64
import re
import zlib
from datetime import datetime

import pandas as pd
import pytest
from openpyxl import load_workbook

from invoice_models import PatientData
from complete_patient_invoice_generator import PatientInvoiceGenerator
from excel_invoice_generator import generate_excel_invoice
from tests.conftest import TEST_CLINIC_CONFIG


def _pdf_text(pdf_bytes: bytes) -> bytes:
    """Decompress a ReportLab PDF's content streams (FlateDecode +
    ASCII85Decode) and concatenate them, so tests can search for real
    rendered text rather than grepping raw (compressed) file bytes."""
    full_text = b""
    for match in re.finditer(rb"stream\r?\n(.*?)endstream", pdf_bytes, re.DOTALL):
        stream_data = match.group(1).strip(b"\r\n")
        if stream_data.endswith(b"~>"):
            stream_data = stream_data[:-2]
        try:
            full_text += zlib.decompress(base64.a85decode(stream_data, adobe=False))
        except (ValueError, zlib.error):
            continue
    return full_text


def _patient():
    return PatientData(prn="1", first_name="Ravil", last_name="Asadullin", dob="",
                        address_line1="1 Main St", address_line2="", city="Testville",
                        state="CA", postal_code="94000")


def _service_df(n_items=9, copay=30):
    rows = [{"visit_date": f"2026-01-{(i % 28) + 1:02d}", "type_of_service": "Psychotherapy",
             "total_amount": copay, "paid": 0, "copay": copay, "previous_balance": 0} for i in range(n_items)]
    return pd.DataFrame(rows)


@pytest.fixture
def qr_clinic(tmp_path):
    """A real, tiny 1x1 JPEG so resolve_qr_image_bytes() has something to load."""
    from PIL import Image
    img_path = tmp_path / "fake_qr.jpg"
    Image.new("RGB", (10, 10), color="white").save(img_path, format="JPEG")
    return {**TEST_CLINIC_CONFIG, "show_qr": True, "qr_image_path": str(img_path)}


@pytest.fixture
def no_qr_clinic():
    return {**TEST_CLINIC_CONFIG, "show_qr": False}


class TestPdfStub:
    def _generate(self, clinic, tmp_path, n_items=9, statement_date="2026-07-17"):
        gen = PatientInvoiceGenerator(amount_due_strategy="auto", statement_date=statement_date,
                                       clinic_config=clinic)
        df = _service_df(n_items)
        lines, total_due, _ = gen._generate_invoice_lines(df)
        out = tmp_path / "invoice.pdf"
        gen._generate_pdf_invoice(_patient(), lines, total_due, df, out)
        return out, total_due, gen

    def test_stub_boxes_all_present_and_before_title(self, no_qr_clinic, tmp_path):
        out, _total_due, _gen = self._generate(no_qr_clinic, tmp_path)
        text = _pdf_text(out.read_bytes())
        for label in [b"STATEMENT DATE:", b"Payment due date:", b"YOUR PORTION DUE:", b"AMOUNT ENCLOSED:"]:
            assert label in text, f"{label!r} missing from stub"
        assert text.index(b"STATEMENT DATE:") < text.index(b"PATIENT STATEMENT")
        assert text.index(b"YOUR PORTION DUE:") < text.index(b"PATIENT STATEMENT")

    def test_amount_due_not_duplicated_at_bottom(self, no_qr_clinic, tmp_path):
        """YOUR PORTION DUE / AMOUNT ENCLOSED must appear exactly once each
        (in the stub) — not also re-rendered below the totals as before."""
        out, _total_due, _gen = self._generate(no_qr_clinic, tmp_path)
        text = _pdf_text(out.read_bytes())
        assert text.count(b"YOUR PORTION DUE:") == 1
        assert text.count(b"AMOUNT ENCLOSED:") == 1

    def test_tear_line_caption_present_and_clean(self, no_qr_clinic, tmp_path):
        out, _total_due, _gen = self._generate(no_qr_clinic, tmp_path)
        text = _pdf_text(out.read_bytes())
        assert b"Detach and retain for your records" in text
        idx = text.index(b"STATEMENT DATE:")
        title_idx = text.index(b"PATIENT STATEMENT")
        tear_idx = text.index(b"Detach and retain for your records")
        assert idx < tear_idx < title_idx

    def test_qr_caption_present_when_enabled(self, qr_clinic, tmp_path):
        out, _total_due, _gen = self._generate(qr_clinic, tmp_path)
        text = _pdf_text(out.read_bytes())
        assert b"Zelle QR Code" in text
        assert b"/Subtype /Image" in out.read_bytes()

    def test_qr_caption_absent_when_disabled(self, no_qr_clinic, tmp_path):
        out, _total_due, _gen = self._generate(no_qr_clinic, tmp_path)
        text = _pdf_text(out.read_bytes())
        assert b"Zelle QR Code" not in text
        assert b"/Subtype /Image" not in out.read_bytes()

    def test_asadullin_sample_fits_one_page(self, qr_clinic, tmp_path):
        """The documented acceptance sample: 9 line items, $270 due."""
        out, total_due, gen = self._generate(qr_clinic, tmp_path, n_items=9)
        assert total_due == 270.0
        assert gen._count_pdf_pages(out.read_bytes()) == 1

    def test_zero_balance_patient_fits_one_page(self, qr_clinic, tmp_path):
        gen = PatientInvoiceGenerator(amount_due_strategy="auto", statement_date="2026-07-17",
                                       clinic_config=qr_clinic)
        df = pd.DataFrame([{"visit_date": "2026-01-15", "type_of_service": "Psychotherapy",
                             "total_amount": 30, "paid": 30, "copay": 0, "previous_balance": 0}])
        lines, raw_total_due, _ = gen._generate_invoice_lines(df)
        total_due = max(0, raw_total_due)
        out = tmp_path / "zero.pdf"
        gen._generate_pdf_invoice(_patient(), lines, total_due, df, out)
        assert gen._count_pdf_pages(out.read_bytes()) == 1

    def test_longer_patient_still_fits_one_page(self, qr_clinic, tmp_path):
        """A patient with many more visits than typical still needs to
        compress down to one page with the QR-enlarged stub in place."""
        out, _total_due, gen = self._generate(qr_clinic, tmp_path, n_items=16)
        assert gen._count_pdf_pages(out.read_bytes()) == 1

    def test_22_item_patient_still_fits_one_page(self, qr_clinic, tmp_path):
        """The QR used to be a flowable cell in the patient/payment table,
        where a fixed-vs-tier-scaled top-padding bug once caused a real
        near-miss with the header at 18 items (see CHANGELOG). It's now
        drawn as an absolutely-positioned image outside the flowable frame
        entirely (add_page_furniture()), so it can no longer inflate that
        row's height or interact with the layout-tier system at all — the
        old failure mode is structurally gone, not just patched. Single-
        page capacity actually grew as a result (up to ~22 items, vs ~18
        with the QR still in the table)."""
        out, _total_due, gen = self._generate(qr_clinic, tmp_path, n_items=22)
        assert gen._count_pdf_pages(out.read_bytes()) == 1

    def test_qr_drawn_via_page_furniture_not_story_table(self, qr_clinic, tmp_path):
        """The QR must be produced by the page-level canvas callback, not
        as content inside the flowable story — guards against the QR
        silently moving back into the patient/payment table (the source of
        the row-height/collision issues fixed by floating it out)."""
        gen = PatientInvoiceGenerator(amount_due_strategy="auto", statement_date="2026-07-17",
                                       clinic_config=qr_clinic)
        assert hasattr(gen, "add_page_furniture")
        assert hasattr(gen, "add_first_page_furniture")
        assert not hasattr(gen, "add_optimized_footer")

    def test_qr_appears_on_page_one_only(self, qr_clinic, tmp_path):
        """A long enough invoice to genuinely overflow to page 2 (the QR
        floats independently of the layout-tier compression, so no amount
        of items keeps this on one page forever) must show the QR/caption
        exactly once — on page 1. It used to be drawn by the same callback
        for every page, so it visually landed on top of whatever
        continuation content (often SUBTOTAL/TOTAL/signature) started at
        the top of page 2."""
        out, _total_due, gen = self._generate(qr_clinic, tmp_path, n_items=30)
        assert gen._count_pdf_pages(out.read_bytes()) == 2
        text = _pdf_text(out.read_bytes())
        assert text.count(b"Zelle QR Code") == 1

    def test_overflow_table_header_repeats_on_page_two(self, qr_clinic, tmp_path):
        """For a genuinely long patient that can't fit one page even at the
        tightest layout tier, the item table's header row must repeat on
        the continuation page rather than leaving page 2 headerless."""
        out, _total_due, gen = self._generate(qr_clinic, tmp_path, n_items=30)
        assert gen._count_pdf_pages(out.read_bytes()) == 2
        text = _pdf_text(out.read_bytes())
        # PDF string literals escape parens, so "Service Date(s)" is stored
        # as "Service Date\(s\)" — search the unambiguous parenless prefix.
        assert text.count(b"Service Date") == 2
        assert text.count(b"Copay/Deductible") == 2


class TestExcelStub:
    def _generate(self, clinic, tmp_path, n_items=9):
        df = _service_df(n_items)
        gen = PatientInvoiceGenerator(amount_due_strategy="auto", statement_date="2026-07-17",
                                       clinic_config=clinic)
        lines, total_due, _ = gen._generate_invoice_lines(df)
        out = tmp_path / "invoice.xlsx"
        generate_excel_invoice(_patient(), lines, total_due, df, gen.statement_date,
                                gen.payment_due_date, False, out, clinic=clinic)
        return out, total_due

    def _cell_row(self, ws, value):
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == value:
                    return cell.row
        return None

    def test_stub_boxes_before_title(self, no_qr_clinic, tmp_path):
        out, _ = self._generate(no_qr_clinic, tmp_path)
        ws = load_workbook(out).active
        stmt_row = self._cell_row(ws, "STATEMENT DATE:")
        due_row = self._cell_row(ws, "YOUR PORTION DUE:")
        title_row = self._cell_row(ws, "PATIENT STATEMENT")
        header_row = self._cell_row(ws, "Service Date(s)")
        assert stmt_row is not None and due_row is not None and title_row is not None
        assert stmt_row < title_row
        assert due_row < title_row < header_row

    def test_amount_due_not_duplicated_at_bottom(self, no_qr_clinic, tmp_path):
        out, _ = self._generate(no_qr_clinic, tmp_path)
        ws = load_workbook(out).active
        count = sum(1 for row in ws.iter_rows() for c in row if c.value == "YOUR PORTION DUE:")
        assert count == 1

    def test_tear_line_row_present(self, no_qr_clinic, tmp_path):
        out, _ = self._generate(no_qr_clinic, tmp_path)
        ws = load_workbook(out).active
        row = self._cell_row(ws, "- - - - -  Detach and retain for your records  - - - - -")
        assert row is not None
        assert ws.cell(row=row, column=1).border.bottom.style == "dashed"

    def test_signature_directly_follows_items_no_old_box(self, no_qr_clinic, tmp_path):
        out, _ = self._generate(no_qr_clinic, tmp_path)
        ws = load_workbook(out).active
        total_row = self._cell_row(ws, "TOTAL")
        sig_row = self._cell_row(ws, "_________________________________")
        # Only a small items-spacer gap between TOTAL and the signature —
        # no room for the old bottom amount box (which was 2 rows + a
        # spacer) to have silently survived in between.
        assert sig_row - total_row <= 3

    def test_qr_enlarged_and_captioned_when_enabled(self, qr_clinic, tmp_path):
        import zipfile
        out, _ = self._generate(qr_clinic, tmp_path)
        ws = load_workbook(out).active
        assert len(ws._images) == 1
        # Column C (index 2): the spacer column between the patient-address
        # block (A:B) and the payment-notice box (D:E) — not the top-right
        # corner, which collides with the merged, full-width, centered
        # clinic header. Anchor row must be at/after the row the patient
        # info block starts on, not up in the header rows.
        assert ws._images[0].anchor._from.col == 2
        stmt_date_row = self._cell_row(ws, "STATEMENT DATE:")
        qr_anchor_row_1indexed = ws._images[0].anchor._from.row + 1
        assert qr_anchor_row_1indexed > 8, "QR must not sit up in the clinic header rows"
        assert qr_anchor_row_1indexed < stmt_date_row, "QR must not sit at/below the stub row"

        caption_row = self._cell_row(ws, "Zelle QR Code")
        assert caption_row is not None
        # Caption lives in column C too, directly beneath the image — not
        # column E, which is inside the payment-instructions merge.
        assert ws.cell(row=caption_row, column=3).value == "Zelle QR Code"

        with zipfile.ZipFile(out) as z:
            drawing_xml = z.read("xl/drawings/drawing1.xml").decode()
        ext_match = re.search(r'<ext cx="(\d+)" cy="(\d+)"', drawing_xml)
        assert ext_match is not None
        cx_inches = int(ext_match.group(1)) / 914400
        assert 1.15 < cx_inches < 1.25, f"expected ~1.2in, got {cx_inches:.3f}in"

    def test_no_qr_image_when_disabled(self, no_qr_clinic, tmp_path):
        out, _ = self._generate(no_qr_clinic, tmp_path)
        ws = load_workbook(out).active
        assert len(ws._images) == 0
        assert self._cell_row(ws, "Zelle QR Code") is None

    def test_asadullin_sample_matches_documented_total(self, qr_clinic, tmp_path):
        out, total_due = self._generate(qr_clinic, tmp_path, n_items=9)
        assert total_due == 270.0
        ws = load_workbook(out).active
        assert ws.sheet_properties.pageSetUpPr.fitToPage is True
