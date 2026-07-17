"""Tests for the optional QR code on invoices (clinic_config: show_qr /
qr_image_path / qr_content). All patient data here is synthetic."""
import pandas as pd
import pytest

from invoice_models import PatientData
from qr_code import generate_qr_png_bytes, qr_settings, resolve_qr_image_bytes
from tests.conftest import TEST_CLINIC_CONFIG


def test_generate_qr_png_bytes_produces_valid_png():
    png = generate_qr_png_bytes("https://example.com/pay")
    assert png[:8] == b"\x89PNG\r\n\x1a\n"  # PNG file signature
    assert len(png) > 100


def test_qr_settings_defaults_to_disabled_when_absent():
    assert qr_settings({}) == (False, "", "")


def test_qr_settings_falls_back_to_website_when_content_missing():
    show, content, image_path = qr_settings({"show_qr": True, "website": "https://clinic.example.com/"})
    assert show is True
    assert content == "https://clinic.example.com/"
    assert image_path == ""


def test_qr_settings_prefers_explicit_qr_content():
    _show, content, _path = qr_settings({
        "show_qr": True, "website": "https://clinic.example.com/",
        "qr_content": "https://clinic.example.com/pay",
    })
    assert content == "https://clinic.example.com/pay"


def test_qr_settings_resolves_qr_image_path():
    _show, _content, image_path = qr_settings({"show_qr": True, "qr_image_path": "templates/zelle_qr.jpg"})
    assert image_path == "templates/zelle_qr.jpg"


class TestResolveQrImageBytes:
    """A static, pre-made QR image (qr_image_path) takes priority over
    generating one from qr_content, since a real payment QR like Zelle's
    encodes bank-specific data a generator can't reproduce from a URL."""

    def test_none_when_disabled(self):
        clinic = {"show_qr": False, "qr_image_path": "templates/zelle_qr.jpg", "qr_content": "https://x.com/pay"}
        assert resolve_qr_image_bytes(clinic) is None

    def test_none_when_nothing_configured(self):
        assert resolve_qr_image_bytes({"show_qr": True}) is None

    def test_static_image_takes_priority(self, tmp_path):
        fake_image = tmp_path / "fake_qr.png"
        fake_image.write_bytes(b"FAKEIMAGEBYTES")
        clinic = {"show_qr": True, "qr_image_path": str(fake_image), "qr_content": "https://x.com/pay"}
        assert resolve_qr_image_bytes(clinic) == b"FAKEIMAGEBYTES"

    def test_falls_back_to_generated_when_image_path_missing_file(self, tmp_path):
        missing_path = tmp_path / "does_not_exist.jpg"
        clinic = {"show_qr": True, "qr_image_path": str(missing_path), "qr_content": "https://x.com/pay"}
        result = resolve_qr_image_bytes(clinic)
        assert result[:8] == b"\x89PNG\r\n\x1a\n"  # generated PNG, not the missing file

    def test_falls_back_to_generated_when_no_image_path(self):
        clinic = {"show_qr": True, "qr_content": "https://x.com/pay"}
        result = resolve_qr_image_bytes(clinic)
        assert result[:8] == b"\x89PNG\r\n\x1a\n"

    def test_none_when_image_path_missing_and_no_qr_content(self, tmp_path):
        missing_path = tmp_path / "does_not_exist.jpg"
        clinic = {"show_qr": True, "qr_image_path": str(missing_path)}
        assert resolve_qr_image_bytes(clinic) is None

    def test_resolves_real_static_template(self):
        """Sanity check against the actual repo asset, so a rename/move of
        templates/zelle_qr.jpg is caught here rather than silently falling
        back to a generated QR."""
        clinic = {"show_qr": True, "qr_image_path": "templates/zelle_qr.jpg"}
        result = resolve_qr_image_bytes(clinic)
        assert result is not None
        assert result[:2] == b"\xff\xd8"  # JPEG file signature


class TestPdfEmbedding:
    def _patient_and_df(self):
        patient = PatientData(prn="1", first_name="Test", last_name="Patient", dob="",
                               address_line1="1 Test St", address_line2="", city="Testville",
                               state="CA", postal_code="94000")
        df = pd.DataFrame([{"visit_date": "2026-01-01", "type_of_service": "Psychotherapy",
                             "total_amount": 200, "paid": 0, "copay": 50, "previous_balance": 0}])
        return patient, df

    def _generate(self, clinic, out_path):
        from complete_patient_invoice_generator import PatientInvoiceGenerator
        gen = PatientInvoiceGenerator(amount_due_strategy="auto", statement_date="2026-07-16",
                                       clinic_config=clinic)
        patient, df = self._patient_and_df()
        lines, total_due, _ = gen._generate_invoice_lines(df)
        gen._generate_pdf_invoice(patient, lines, total_due, df, out_path)

    def test_static_image_embedded_in_pdf_when_enabled(self, tmp_path):
        clinic = {**TEST_CLINIC_CONFIG, "show_qr": True, "qr_image_path": "templates/zelle_qr.jpg"}
        out_path = tmp_path / "qr_static.pdf"
        self._generate(clinic, out_path)
        assert b"/Subtype /Image" in out_path.read_bytes()

    def test_generated_qr_embedded_in_pdf_when_no_static_image(self, tmp_path):
        clinic = {**TEST_CLINIC_CONFIG, "show_qr": True, "qr_image_path": "", "qr_content": "https://example.com/pay"}
        out_path = tmp_path / "qr_generated.pdf"
        self._generate(clinic, out_path)
        assert b"/Subtype /Image" in out_path.read_bytes()

    def test_qr_absent_from_pdf_when_disabled(self, tmp_path):
        clinic = {**TEST_CLINIC_CONFIG, "show_qr": False}
        out_path = tmp_path / "qr_off.pdf"
        self._generate(clinic, out_path)
        assert b"/Subtype /Image" not in out_path.read_bytes()


class TestExcelEmbedding:
    def _patient_and_df(self):
        patient = PatientData(prn="1", first_name="Test", last_name="Patient", dob="",
                               address_line1="1 Test St", address_line2="", city="Testville",
                               state="CA", postal_code="94000")
        df = pd.DataFrame([{"visit_date": "2026-01-01", "type_of_service": "Psychotherapy",
                             "paid": 0, "copay": 50, "previous_balance": 0}])
        return patient, df

    def _generate(self, clinic, out_path):
        from datetime import datetime
        from excel_invoice_generator import generate_excel_invoice
        patient, df = self._patient_and_df()
        generate_excel_invoice(patient=patient, lines=[], total_due=50.0, patient_df=df,
                                statement_date=datetime(2026, 7, 16), payment_due_date=datetime(2026, 8, 17),
                                has_cpt=False, output_path=out_path, clinic=clinic)

    def test_static_image_embedded_in_excel_when_enabled(self, tmp_path):
        from openpyxl import load_workbook
        clinic = {**TEST_CLINIC_CONFIG, "show_qr": True, "qr_image_path": "templates/zelle_qr.jpg"}
        out_path = tmp_path / "qr_static.xlsx"
        self._generate(clinic, out_path)
        ws = load_workbook(out_path).active
        assert len(ws._images) == 1
        # Anchored in column C (index 2), the deliberate spacer between the
        # patient and payment-notice boxes — never overlaps existing content.
        assert ws._images[0].anchor._from.col == 2

    def test_generated_qr_embedded_in_excel_when_no_static_image(self, tmp_path):
        from openpyxl import load_workbook
        clinic = {**TEST_CLINIC_CONFIG, "show_qr": True, "qr_image_path": "", "qr_content": "https://example.com/pay"}
        out_path = tmp_path / "qr_generated.xlsx"
        self._generate(clinic, out_path)
        ws = load_workbook(out_path).active
        assert len(ws._images) == 1

    def test_qr_absent_from_excel_when_disabled(self, tmp_path):
        from openpyxl import load_workbook
        clinic = {**TEST_CLINIC_CONFIG, "show_qr": False}
        out_path = tmp_path / "qr_off.xlsx"
        self._generate(clinic, out_path)
        ws = load_workbook(out_path).active
        assert len(ws._images) == 0

    def test_golden_fixture_unaffected_since_test_config_disables_qr(self):
        """The Phase 0 golden-fixture test config (clinic_config.example.json,
        loaded as TEST_CLINIC_CONFIG) must have show_qr off, or every
        excel_invoice_generator test that doesn't override it would
        silently start embedding an image the golden fixture doesn't have."""
        assert TEST_CLINIC_CONFIG.get("show_qr") is False
